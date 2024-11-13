try:
    import wmi
    import os
    import platform
    import psutil
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    import win32com.client
    import tkinter as tk
    from tkinter import messagebox
except ImportError as e:
    print("Gerekli paketler eksik. Lütfen önce 'npm run setup' komutunu çalıştırın.")
    print(f"Hata: {str(e)}")
    exit(1)

def get_system_info():
    try:
        c = wmi.WMI()
        
        # Bilgisayar sistem bilgileri
        computer_system = c.Win32_ComputerSystem()[0]
        bios = c.Win32_BIOS()[0]
        os_info = c.Win32_OperatingSystem()[0]
        processor = c.Win32_Processor()[0]
        
        # Disk bilgileri
        disk_drive = c.Win32_DiskDrive()[0]
        
        # GPU bilgileri
        gpu = c.Win32_VideoController()[0]
        
        # Network adaptörleri
        network_adapters = c.Win32_NetworkAdapter(PhysicalAdapter=True)
        wifi_mac = ""
        ethernet_mac = ""
        
        for adapter in network_adapters:
            if adapter.MACAddress:
                if "Wireless" in adapter.Name or "Wi-Fi" in adapter.Name:
                    wifi_mac = adapter.MACAddress
                elif "Ethernet" in adapter.Name:
                    ethernet_mac = adapter.MACAddress
        
        # RAM hesaplama
        total_ram = round(float(os_info.TotalVisibleMemorySize) / 1024 / 1024, 2)
        
        # İşletim sistemi bilgilerini ayır
        os_name = os_info.Caption.strip()
        os_version = os_info.Version.strip()
        
        # Sistem türü (32/64 bit)
        system_type = os_info.OSArchitecture
        
        system_info = {
            "Cihaz Adı": computer_system.Name,
            "Marka": computer_system.Manufacturer,
            "Model": computer_system.Model,
            "Seri Numarası": bios.SerialNumber,
            "Wifi MAC": wifi_mac if wifi_mac else "Bulunamadı",
            "Ethernet MAC": ethernet_mac if ethernet_mac else "Bulunamadı",
            "İşletim Sistemi Adı": os_name,
            "İşletim Sistemi Sürümü": os_version,
            "Sistem Türü": system_type,
            "CPU Markası": processor.Manufacturer,
            "CPU Modeli": processor.Name,
            "RAM (GB)": f"{total_ram} GB",
            "Disk Modeli": disk_drive.Model,
            "Disk Kapasitesi": f"{round(float(disk_drive.Size) / 1024 / 1024 / 1024, 2)} GB",
            "GPU Markası": f"{gpu.Name}"
        }
        
        return system_info
    except Exception as e:
        print(f"Sistem bilgileri toplanırken hata oluştu: {str(e)}")
        return None

def save_to_excel(system_info):
    try:
        # Belgelerim klasörünü al
        documents_path = os.path.join(os.path.expanduser('~'), 'Documents')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = os.path.join(documents_path, f"sistem_bilgileri_{timestamp}.xlsx")
        
        # Excel dosyası oluştur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sistem Bilgileri"
        
        # Başlıkları ekle
        headers = list(system_info.keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Verileri ekle
        for col, value in enumerate(system_info.values(), 1):
            ws.cell(row=2, column=col, value=value)
        
        # Sütun genişliklerini ayarla
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Dosyayı kaydet
        wb.save(excel_path)
        print(f"Sistem bilgileri başarıyla kaydedildi: {excel_path}")

        # Mesaj kutusunu göster
        root = tk.Tk()
        root.withdraw()  # Ana pencereyi gizle
        messagebox.showinfo("İşlem Tamamlandı", f"İşlem başarılı bir şekilde tamamlanmıştır. Dosyanız '{excel_path}' buradadır. Kolay gelsin")
        
    except Exception as e:
        print(f"Excel dosyası oluşturulurken hata oluştu: {str(e)}")

def main():
    print("Sistem bilgileri toplanıyor...")
    system_info = get_system_info()
    
    if system_info:
        save_to_excel(system_info)
    else:
        print("Sistem bilgileri toplanamadı!")

if __name__ == "__main__":
    main()
